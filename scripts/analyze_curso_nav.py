# -*- coding: utf-8 -*-
"""Analiza navegacion interactiva en Curso.xlsx"""
import zipfile
import re
from openpyxl import load_workbook

ORIG = r"c:\Proyectos\Aplicacion_Personal\content\excel\Curso.xlsx"
EDIT = r"c:\Proyectos\Aplicacion_Personal\content\excel\Curso_Editable.xlsx"

def analyze(path, label):
    print(f"\n{'='*60}\n{label}: {path}\n{'='*60}")
    wb = load_workbook(path, data_only=False)
    print("Sheets:", wb.sheetnames)
    for name in wb.sheetnames[:5]:
        ws = wb[name]
        print(f"\n--- {name} state={ws.sheet_state} ---")
        links = []
        for row in ws.iter_rows(max_row=15, max_col=20):
            for cell in row:
                if cell.hyperlink:
                    links.append((cell.coordinate, str(cell.value)[:40], cell.hyperlink.target or cell.hyperlink.location))
        if links:
            for l in links[:20]:
                print(f"  LINK {l[0]}: {l[1]} -> {l[2]}")
        # merged cells in row 1-3
        merges = [str(m) for m in ws.merged_cells.ranges if m.min_row <= 3]
        if merges:
            print(f"  Merges row1-3: {merges[:8]}")
    # hyperlinks in workbook rels
    with zipfile.ZipFile(path) as z:
        for n in z.namelist():
            if "worksheets/_rels/sheet" in n and n.endswith(".rels"):
                data = z.read(n).decode("utf-8", errors="replace")
                if "hyperlink" in data.lower():
                    print(f"\n  RELS {n}:")
                    for m in re.finditer(r'Target="([^"]*)"', data):
                        t = m.group(1)
                        if "hyperlink" in t or "#" in t or ".xml" not in t:
                            print(f"    {t}")
    wb.close()

analyze(ORIG, "ORIGINAL")
if __import__("os").path.exists(EDIT):
    analyze(EDIT, "EDITABLE")
