# -*- coding: utf-8 -*-
import re
import zipfile
from openpyxl import load_workbook

P = r"c:\Proyectos\Aplicacion_Personal\content\excel\Curso.xlsx"

wb = load_workbook(P, read_only=True)
# sheet id mapping from workbook.xml
with zipfile.ZipFile(P) as z:
    wb_xml = z.read("xl/workbook.xml").decode("utf-8")
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    # rId to sheet file
    rid_map = {}
    for m in re.finditer(r'Id="(rId\d+)"[^>]*Target="worksheets/(sheet\d+\.xml)"', rels):
        rid_map[m.group(1)] = m.group(2)

    sheets_info = []
    for m in re.finditer(
        r'<sheet[^>]*name="([^"]*)"[^>]*r:id="(rId\d+)"[^>]*/>', wb_xml
    ):
        name, rid = m.group(1), m.group(2)
        file = rid_map.get(rid, "?")
        sheets_info.append((name, file))

    print("=== PROTECCION POR HOJA (sheet=1 = bloqueada con password) ===\n")
    for name, file in sheets_info:
        if file == "?":
            continue
        path = f"xl/worksheets/{file}"
        t = z.read(path).decode("utf-8", errors="replace")
        tag_m = re.search(r"<sheetProtection([^/]*)/>", t)
        if not tag_m:
            print(f"{name}: sin proteccion")
            continue
        attrs = tag_m.group(1)
        active = "sheet=\"1\"" in attrs or "sheet='1'" in attrs
        has_pwd = "hashValue" in attrs
        state = wb[name].sheet_state if name in wb.sheetnames else "?"
        if active or has_pwd:
            print(f"{name} [{state}] ({file}): ACTIVA password={has_pwd} attrs={attrs[:100]}...")
        else:
            print(f"{name} [{state}]: proteccion XML inactiva (sheet=0)")

wb.close()
