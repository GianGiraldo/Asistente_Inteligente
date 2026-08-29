# -*- coding: utf-8 -*-
"""Genera Curso_Gestion_Empresarial_Andina.xlsx con demos funcionales del curso."""
import os
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

OUTPUT = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "content",
    "excel",
    "Curso_Gestion_Empresarial_Andina.xlsx",
)

# Paleta corporativa Distribuidora Andina
C_PRIMARY = "00B4D8"
C_DARK = "1A1A2E"
C_POS = "2ECC71"
C_NEG = "E74C3C"
C_WARN = "F39C12"
C_BG = "F5F7FA"
C_GRAY = "95A5A6"

THIN = Side(style="thin", color="D0D5DD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def sty_header(cell, bg=C_DARK, fg="FFFFFF"):
    cell.font = Font(bold=True, color=fg, name="Segoe UI", size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def sty_kpi(cell, size=22):
    cell.font = Font(bold=True, color=C_DARK, name="Segoe UI", size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def sty_title(ws, text, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=14, color=C_DARK, name="Segoe UI")
    c.fill = PatternFill("solid", fgColor=C_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")


def add_table(ws, name, ref, style="TableStyleMedium2"):
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def build_index(wb):
    ws = wb.active
    ws.title = "00_Indice"
    ws.sheet_properties.tabColor = C_PRIMARY
    sty_title(ws, "CURSO AVANZADO — Gestión Empresarial con Excel | Distribuidora Andina SAC", 1)
    ws.cell(2, 1, value="Nivel Avanzado · Excel 365/2021 · veloX Minicursos")
    ws.cell(2, 1).font = Font(color=C_GRAY, name="Segoe UI", size=10)

    headers = ["Módulo", "Hoja demo", "Funcionalidad activa"]
    for i, h in enumerate(headers, 1):
        sty_header(ws.cell(4, i, h))

    rows = [
        ("1", "01_Inventario_Alertas", "Tabla oficial + formato condicional + fórmulas cobertura"),
        ("1", "02_Facturas_Vencimiento", "Formato condicional fila completa (vencidas/alerta/pagado)"),
        ("1", "03_Catalogo_Sucursales", "Catálogo maestro (base validación cascada)"),
        ("1", "04_Formulario_Cascada", "Validación dependiente Región → Sucursal"),
        ("2", "05_Ventas_Cruce", "BUSCARX / segmento cliente + margen S/"),
        ("2", "06_Maestro_Clientes", "Maestro clientes (dimensión)"),
        ("2", "07_Maestro_Productos", "Maestro productos"),
        ("2", "08_Liquidacion_SUMIFS", "SUMAR.SI.CONJUNTO comisiones por vendedor"),
        ("2", "09_Cartera_Mora", "DIAS.LAB.INTL + clasificación SI.CONJUNTO"),
        ("3", "10_Datos_Pivot_Ventas", "Datos listos para Tabla Dinámica (Insertar → Pivot)"),
        ("5", "11_Dashboard_KPIs", "KPIs ejecutivos + gráfico + semáforos"),
        ("7", "12_Simulador_Meta", "Buscar objetivo — modelo utilidad"),
        ("7", "13_Escenarios", "Variables escenario Base / Crisis / Optimista"),
        ("7", "14_Solver_Promociones", "Modelo listo para Solver (maximizar utilidad)"),
        ("6", "15_Automatizacion_VBA", "Código VBA de referencia (copiar al Editor)"),
        ("—", "16_Guia_Rapida", "Resumen 7 módulos del curso"),
        ("—", "17_Mapas_Contenido_Curso", "22 sub-temas: gancho + fórmula + enlace demo"),
    ]
    for r, row in enumerate(rows, 5):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = BORDER
        link = f"#'{row[1]}'!A1"
        ws.cell(r, 2).hyperlink = link
        ws.cell(r, 2).font = Font(color=C_PRIMARY, underline="single", name="Segoe UI")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 52


def build_inventario(wb):
    ws = wb.create_sheet("01_Inventario_Alertas")
    ws.sheet_properties.tabColor = C_POS
    sty_title(ws, "Módulo 1 · Inventario con alertas automáticas (Formato condicional + Tabla Oficial)")

    headers = [
        "SKU", "Producto", "Stock", "Margen %", "Venta diaria",
        "Lead time (días)", "Días cobertura", "Semáforo",
    ]
    for i, h in enumerate(headers, 1):
        sty_header(ws.cell(3, i, h))

    data = [
        ("AND-001", "Arroz 5kg", 45, 8.2, 12, 5),
        ("AND-042", "Aceite 1L", 8, -2.1, 6, 7),
        ("AND-088", "Detergente", 120, 15.0, 4, 3),
        ("AND-105", "Leche entera 1L", 22, 6.5, 18, 4),
        ("AND-120", "Galletas surtidas", 95, 11.3, 8, 5),
        ("AND-156", "Agua 2L", 200, 9.8, 25, 3),
        ("AND-201", "Shampoo 400ml", 14, -1.5, 3, 10),
        ("AND-215", "Café 250g", 38, 14.2, 5, 6),
    ]
    start = 4
    for r, row in enumerate(data, start):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
        # Días cobertura
        ws.cell(r, 7, value=f"=SI(E{r}>0;C{r}/E{r};999)")
        # Semáforo
        ws.cell(r, 8, value=f'=SI(G{r}>=F{r}*1,5;"VERDE";SI(G{r}>=F{r};"AMARILLO";"ROJO"))')

    end = start + len(data) - 1
    add_table(ws, "tblInventario", f"A3:H{end}")

    # Formato condicional: fila crítica stock
    red_fill = PatternFill("solid", fgColor=C_NEG)
    white_font = Font(color="FFFFFF", bold=True)
    ws.conditional_formatting.add(
        f"A{start}:H{end}",
        FormulaRule(formula=[f"Y(G{start}<F{start};G{start}<999)"], fill=red_fill, font=white_font),
    )
    # Margen negativo columna D
    ws.conditional_formatting.add(
        f"D{start}:D{end}",
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FFC7CE")),
    )
    # Semáforo colores
    ws.conditional_formatting.add(
        f"H{start}:H{end}",
        FormulaRule(formula=[f'H{start}="ROJO"'], fill=red_fill, font=white_font),
    )
    ws.conditional_formatting.add(
        f"H{start}:H{end}",
        FormulaRule(
            formula=[f'H{start}="AMARILLO"'],
            fill=PatternFill("solid", fgColor=C_WARN),
        ),
    )
    ws.conditional_formatting.add(
        f"H{start}:H{end}",
        FormulaRule(
            formula=[f'H{start}="VERDE"'],
            fill=PatternFill("solid", fgColor=C_POS),
            font=Font(color="FFFFFF", bold=True),
        ),
    )

    ws.cell(end + 2, 1, value="NOTA: AND-042 → cobertura crítica + margen negativo (orden urgente 200 u.)")
    ws.cell(end + 2, 1).font = Font(bold=True, color=C_NEG, name="Segoe UI")

    for col, w in enumerate([12, 22, 8, 10, 12, 14, 14, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_facturas(wb):
    ws = wb.create_sheet("02_Facturas_Vencimiento")
    ws.sheet_properties.tabColor = C_WARN
    sty_title(ws, "Módulo 1 · Facturas — colores automáticos según vencimiento")

    headers = ["Factura", "Fecha Venc.", "Estado", "Cliente", "Monto S/"]
    for i, h in enumerate(headers, 1):
        sty_header(ws.cell(3, i, h))

    today = date.today()
    facturas = [
        ("F-8821", today - timedelta(days=12), "Pendiente", "Bodega El Sol", 12400),
        ("F-8902", today - timedelta(days=3), "Pendiente", "Minimarket 24", 8900),
        ("F-8955", today + timedelta(days=5), "Pendiente", "Rest. La Casona", 5600),
        ("F-9010", today + timedelta(days=2), "Pendiente", "Hotel Plaza", 22100),
        ("F-9033", today - timedelta(days=30), "Pagado", "Distrib. Norte", 45000),
        ("F-9088", today + timedelta(days=15), "Pendiente", "Super Andino", 18750),
        ("F-9112", today - timedelta(days=1), "Pendiente", "Cafetería Central", 3200),
    ]
    start = 4
    for r, row in enumerate(facturas, start):
        ws.cell(r, 1, row[0])
        ws.cell(r, 2, row[1])
        ws.cell(r, 2).number_format = "dd/mm/yyyy"
        ws.cell(r, 3, row[2])
        ws.cell(r, 4, row[3])
        ws.cell(r, 5, row[4])
        ws.cell(r, 5).number_format = "#,##0"

    end = start + len(facturas) - 1
    add_table(ws, "tblFacturas", f"A3:E{end}")

    green = PatternFill("solid", fgColor=C_POS)
    yellow = PatternFill("solid", fgColor=C_WARN)
    red = PatternFill("solid", fgColor=C_NEG)
    wfont = Font(color="FFFFFF")

    ws.conditional_formatting.add(
        f"A{start}:E{end}",
        FormulaRule(formula=[f"C{start}=\"Pagado\""], fill=green, font=Font(color="FFFFFF", bold=True)),
    )
    ws.conditional_formatting.add(
        f"A{start}:E{end}",
        FormulaRule(
            formula=[f"Y(C{start}<>\"Pagado\";B{start}>=HOY();B{start}<=HOY()+7)"],
            fill=yellow,
        ),
    )
    ws.conditional_formatting.add(
        f"A{start}:E{end}",
        FormulaRule(
            formula=[f"Y(C{start}<>\"Pagado\";B{start}<HOY())"],
            fill=red,
            font=wfont,
        ),
    )

    for col, w in enumerate([12, 14, 12, 24, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_catalogo(wb):
    ws = wb.create_sheet("03_Catalogo_Sucursales")
    ws.sheet_properties.tabColor = C_GRAY
    sty_title(ws, "Catálogo maestro — base para validación en cascada")

    headers = ["Región", "Sucursal", "Almacén"]
    for i, h in enumerate(headers, 1):
        sty_header(ws.cell(3, i, h))

    catalog = [
        ("Lima", "Miraflores", "ALM-LIM-01"),
        ("Lima", "San Juan", "ALM-LIM-02"),
        ("Lima", "Comas", "ALM-LIM-03"),
        ("Norte", "Trujillo", "ALM-NOR-01"),
        ("Norte", "Chiclayo", "ALM-NOR-02"),
        ("Sur", "Arequipa", "ALM-SUR-01"),
        ("Sur", "Cusco", "ALM-SUR-02"),
        ("Centro", "Huancayo", "ALM-CEN-01"),
    ]
    for r, row in enumerate(catalog, 4):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)

    add_table(ws, "tblSucursales", f"A3:C{3 + len(catalog)}")

    # Named ranges per region for INDIRECT fallback
    regions = {"Lima": [], "Norte": [], "Sur": [], "Centro": []}
    for reg, suc, _ in catalog:
        regions[reg].append(suc)

    for reg, sucs in regions.items():
        # Store lists in columns E+ for named ranges
        col_idx = 5 + list(regions.keys()).index(reg)
        ws.cell(3, col_idx, reg)
        sty_header(ws.cell(3, col_idx))
        for i, s in enumerate(sucs, 4):
            ws.cell(i, col_idx, s)
        ref = f"'03_Catalogo_Sucursales'!${get_column_letter(col_idx)}$4:${get_column_letter(col_idx)}${3 + len(sucs)}"
        from openpyxl.workbook.defined_name import DefinedName

        col_letter = get_column_letter(col_idx)
        end_row = 3 + len(sucs)
        attr = f"'03_Catalogo_Sucursales'!${col_letter}$4:${col_letter}${end_row}"
        wb.defined_names.add(DefinedName(reg, attr_text=attr))

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14


def build_formulario(wb):
    ws = wb.create_sheet("04_Formulario_Cascada")
    ws.sheet_properties.tabColor = C_PRIMARY
    sty_title(ws, "Módulo 1 · Formulario — Región → Sucursal (validación dependiente)")

    ws.cell(5, 2, "Región:")
    ws.cell(5, 2).font = Font(bold=True, name="Segoe UI")
    ws.cell(7, 2, "Sucursal:")
    ws.cell(7, 2).font = Font(bold=True, name="Segoe UI")
    ws.cell(9, 2, "Almacén:")
    ws.cell(9, 2).font = Font(bold=True, name="Segoe UI")
    ws.cell(11, 2, "SKU:")
    ws.cell(11, 2).font = Font(bold=True, name="Segoe UI")
    ws.cell(13, 2, "Cantidad:")
    ws.cell(13, 2).font = Font(bold=True, name="Segoe UI")

    # Region dropdown
    dv_reg = DataValidation(type="list", formula1='"Lima,Norte,Sur,Centro"', allow_blank=False)
    ws.add_data_validation(dv_reg)
    dv_reg.add(ws["C6"])

    # Sucursal INDIRECT
    dv_suc = DataValidation(type="list", formula1="=INDIRECTO(C6)", allow_blank=False)
    ws.add_data_validation(dv_suc)
    dv_suc.add(ws["C8"])

    # Almacén via BUSCARX
    ws.cell(10, 3, value="=BUSCARX(C8;tblSucursales[Sucursal];tblSucursales[Almacén];\"N/A\")")

    ws.cell(12, 3, value="AND-001")
    ws.cell(14, 3, value=50)

    ws.cell(16, 2, value="Instrucción: Cambia Región en C6 → el dropdown de Sucursal en C8 se actualiza.")
    ws.cell(16, 2).font = Font(color=C_PRIMARY, name="Segoe UI", italic=True)

    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 24


def build_maestro_clientes(wb):
    ws = wb.create_sheet("06_Maestro_Clientes")
    ws.sheet_properties.tabColor = "3498DB"
    sty_title(ws, "Maestro Clientes — dimensión para BUSCARX / modelo de datos")

    headers = ["RUC", "Nombre", "Segmento", "Ciudad"]
    for i, h in enumerate(headers, 1):
        sty_header(ws.cell(3, i, h))

    clients = [
        ("20123456789", "Bodega El Sol", "A", "Trujillo"),
        ("20999888777", "Minimarket 24", "B", "Lima"),
        ("20444555666", "Super Andino", "A", "Arequipa"),
        ("20555666777", "Hotel Plaza", "A", "Lima"),
        ("20666777888", "Rest. La Casona", "B", "Cusco"),
        ("20777888999", "Distrib. Norte", "C", "Chiclayo"),
    ]
    for r, row in enumerate(clients, 4):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)

    add_table(ws, "tblClientes", f"A3:D{3 + len(clients)}")
    for col, w in enumerate([14, 24, 10, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_maestro_productos(wb):
    ws = wb.create_sheet("07_Maestro_Productos")
    ws.sheet_properties.tabColor = "3498DB"
    sty_title(ws, "Maestro Productos")

    headers = ["SKU", "Nombre", "Categoría", "Costo S/"]
    for i, h in enumerate(headers, 1):
        sty_header(ws.cell(3, i, h))

    products = [
        ("AND-001", "Arroz 5kg", "Abarrotes", 18.50),
        ("AND-042", "Aceite 1L", "Abarrotes", 9.80),
        ("AND-088", "Detergente", "Limpieza", 6.20),
        ("AND-105", "Leche 1L", "Lácteos", 4.10),
        ("AND-120", "Galletas", "Snacks", 3.50),
    ]
    for r, row in enumerate(products, 4):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
        ws.cell(r, 4).number_format = "#,##0.00"

    add_table(ws, "tblProductos", f"A3:D{3 + len(products)}")
    for col, w in enumerate([12, 22, 12, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_ventas(wb):
    ws = wb.create_sheet("05_Ventas_Cruce")
    ws.sheet_properties.tabColor = "9B59B6"
    sty_title(ws, "Módulo 2 · Ventas con BUSCARX (segmento + categoría producto)")

    headers = [
        "ID", "RUC", "SKU", "Monto S/", "Margen %", "Vendedor", "Mes",
        "Segmento", "Categoría", "Margen S/",
    ]
    for i, h in enumerate(headers, 1):
        sty_header(ws.cell(3, i, h))

    ventas = [
        ("V-001", "20123456789", "AND-001", 4500, 8.2, "Ana Torres", "Mayo"),
        ("V-002", "20999888777", "AND-042", 890, 5.1, "Luis Mendoza", "Mayo"),
        ("V-003", "20444555666", "AND-088", 3200, 15.0, "Ana Torres", "Mayo"),
        ("V-004", "20555666777", "AND-105", 22100, 6.5, "Carla Ríos", "Mayo"),
        ("V-005", "20666777888", "AND-120", 5600, 11.3, "Luis Mendoza", "Mayo"),
        ("V-006", "20777888999", "AND-001", 12500, 8.2, "Pedro Sánchez", "Mayo"),
        ("V-007", "20123456789", "AND-088", 2100, 15.0, "Ana Torres", "Mayo"),
        ("V-008", "20999888777", "AND-105", 4400, 6.5, "Luis Mendoza", "Mayo"),
    ]
    start = 4
    for r, row in enumerate(ventas, start):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
        ws.cell(r, 4).number_format = "#,##0"
        ws.cell(r, 8, value=f'=BUSCARX(B{r};tblClientes[RUC];tblClientes[Segmento];"REVISAR")')
        ws.cell(r, 9, value=f'=BUSCARX(C{r};tblProductos[SKU];tblProductos[Categoría];"N/A")')
        ws.cell(r, 10, value=f"=D{r}*E{r}/100")
        ws.cell(r, 10).number_format = "#,##0.00"

    end = start + len(ventas) - 1
    add_table(ws, "tblVentas", f"A3:J{end}")

    ws.cell(end + 2, 1, value="REVISAR en Segmento = RUC sin maestro cliente")
    ws.cell(end + 2, 1).font = Font(color=C_WARN, name="Segoe UI")

    for col, w in enumerate([8, 14, 12, 10, 10, 14, 8, 10, 12, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_liquidacion(wb):
    ws = wb.create_sheet("08_Liquidacion_SUMIFS")
    ws.sheet_properties.tabColor = "9B59B6"
    sty_title(ws, "Módulo 2 · Liquidación comisiones — SUMAR.SI.CONJUNTO")

    headers = ["Vendedor", "Ventas Mayo S/", "Margen S/ Mayo", "Pedidos Mayo"]
    for i, h in enumerate(headers, 1):
        sty_header(ws.cell(3, i, h))

    vendedores = ["Ana Torres", "Luis Mendoza", "Carla Ríos", "Pedro Sánchez"]
    start = 4
    for r, vend in enumerate(vendedores, start):
        ws.cell(r, 1, vend)
        ws.cell(r, 2, value=f'=SUMAR.SI.CONJUNTO(tblVentas[Monto S/];tblVentas[Vendedor];A{r};tblVentas[Mes];"Mayo")')
        ws.cell(r, 2).number_format = "#,##0"
        ws.cell(r, 3, value=f'=SUMAR.SI.CONJUNTO(tblVentas[Margen S/];tblVentas[Vendedor];A{r};tblVentas[Mes];"Mayo")')
        ws.cell(r, 3).number_format = "#,##0.00"
        ws.cell(r, 4, value=f'=CONTAR.SI.CONJUNTO(tblVentas[ID];tblVentas[Vendedor];A{r};tblVentas[Mes];"Mayo")')

    ws.cell(9, 1, value="Total equipo:")
    ws.cell(9, 1).font = Font(bold=True)
    ws.cell(9, 2, value="=SUMA(B4:B7)")
    ws.cell(9, 2).number_format = "#,##0"
    ws.cell(9, 3, value="=SUMA(C4:C7)")
    ws.cell(9, 3).number_format = "#,##0.00"

    for col, w in enumerate([16, 16, 16, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_cartera(wb):
    today = date.today()
    ws = wb.create_sheet("09_Cartera_Mora")
    ws.sheet_properties.tabColor = C_NEG
    sty_title(ws, "Módulo 2 · Cartera — DIAS.LAB.INTL + clasificación")

    # Feriados en la misma hoja (rango tblFeriados)
    ws.cell(3, 7, value="Feriados")
    sty_header(ws.cell(3, 7))
    feriados = [
        date(2024, 1, 1), date(2024, 7, 28), date(2024, 12, 25),
        date(2025, 1, 1), date(2025, 7, 28), date(2025, 12, 25),
    ]
    for i, d in enumerate(feriados, 4):
        ws.cell(i, 7, d)
        ws.cell(i, 7).number_format = "dd/mm/yyyy"
    from openpyxl.workbook.defined_name import DefinedName

    wb.defined_names.add(
        DefinedName("tblFeriados", attr_text="'09_Cartera_Mora'!$G$4:$G$9")
    )

    headers = ["Factura", "Vencimiento", "Monto S/", "Mora hábil (días)", "Clasificación"]
    for i, h in enumerate(headers, 1):
        sty_header(ws.cell(3, i, h))

    cartera = [
        ("F-8821", today - timedelta(days=18), 12400),
        ("F-8902", today - timedelta(days=8), 8900),
        ("F-8955", today + timedelta(days=5), 5600),
        ("F-9010", today - timedelta(days=25), 22100),
        ("F-9112", today - timedelta(days=14), 3200),
    ]
    today_str = "HOY()"
    start = 4
    for r, row in enumerate(cartera, start):
        ws.cell(r, 1, row[0])
        ws.cell(r, 2, row[1])
        ws.cell(r, 2).number_format = "dd/mm/yyyy"
        ws.cell(r, 3, row[2])
        ws.cell(r, 3).number_format = "#,##0"
        ws.cell(r, 4, value=f'=SI(B{r}<HOY();DIAS.LAB.INTL(B{r};HOY();11;tblFeriados);0)')
        ws.cell(r, 5, value=f'=SI.CONJUNTO(D{r}>10;"GESTIÓN INTENSA";D{r}>5;"SEGUIMIENTO";D{r}>0;"RECORDATORIO";VERDADERO;"AL DÍA")')

    end = start + len(cartera) - 1
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        FormulaRule(formula=[f'E{start}="GESTIÓN INTENSA"'], fill=PatternFill("solid", fgColor=C_NEG), font=Font(color="FFFFFF", bold=True)),
    )

    for col, w in enumerate([12, 14, 12, 16, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_pivot_data(wb):
    ws = wb.create_sheet("10_Datos_Pivot_Ventas")
    ws.sheet_properties.tabColor = "E67E22"
    sty_title(ws, "Módulo 3 · Datos para Tabla Dinámica — Insertar → Tabla dinámica desde tblVentasPivot")

    headers = ["Fecha", "Sucursal", "Región", "Categoría", "Vendedor", "Monto S/", "Margen %", "Mes"]
    for i, h in enumerate(headers, 1):
        sty_header(ws.cell(3, i, h))

    sucursales = [
        ("Miraflores", "Lima"), ("Trujillo", "Norte"), ("Arequipa", "Sur"),
        ("Huancayo", "Centro"), ("Chiclayo", "Norte"),
    ]
    cats = ["Abarrotes", "Lácteos", "Limpieza", "Snacks"]
    vends = ["Ana Torres", "Luis Mendoza", "Carla Ríos"]
    rows = []
    base = date(2024, 5, 1)
    for i in range(40):
        suc, reg = sucursales[i % 5]
        rows.append((
            base + timedelta(days=i % 28),
            suc, reg, cats[i % 4], vends[i % 3],
            1500 + (i * 317) % 8000,
            5 + (i % 12),
            "Mayo",
        ))

    start = 4
    for r, row in enumerate(rows, start):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            if c == 1:
                cell.number_format = "dd/mm/yyyy"
            if c == 6:
                cell.number_format = "#,##0"

    end = start + len(rows) - 1
    add_table(ws, "tblVentasPivot", f"A3:H{end}")

    ws.cell(end + 2, 1, value="PIVOT: Región × Categoría × Suma Monto | Segmentar: Mes, Sucursal")
    ws.cell(end + 2, 1).font = Font(color=C_PRIMARY, name="Segoe UI", bold=True)

    for col, w in enumerate([12, 14, 10, 12, 14, 12, 10, 8], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_dashboard(wb):
    ws = wb.create_sheet("11_Dashboard_KPIs")
    ws.sheet_properties.tabColor = C_DARK
    ws.sheet_view.showGridLines = False

    # Header bar
    for c in range(1, 9):
        ws.cell(1, c).fill = PatternFill("solid", fgColor=C_DARK)
    ws.merge_cells("A1:H1")
    ws.cell(1, 1, value="DISTRIBUIDORA ANDINA SAC — Dashboard Ejecutivo")
    ws.cell(1, 1).font = Font(bold=True, size=16, color="FFFFFF", name="Segoe UI")
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(2, 1, value="=TEXTO(AHORA();\"Actualizado: dd/mm/yyyy hh:mm\")")
    ws.cell(2, 1).font = Font(color=C_GRAY, name="Segoe UI", size=9)

    # KPI cards
    kpis = [
        ("Ventas Mayo S/", "=SUMAR.SI.CONJUNTO(tblVentas[Monto S/];tblVentas[Mes];\"Mayo\")", C_PRIMARY),
        ("Margen S/ Mayo", "=SUMAR.SI.CONJUNTO(tblVentas[Margen S/];tblVentas[Mes];\"Mayo\")", C_POS),
        ("Pedidos Mayo", "=CONTAR.SI(tblVentas[Mes];\"Mayo\")", C_DARK),
        ("Ticket promedio", "=SIERROR(B6/B8;0)", C_PRIMARY),
    ]
    positions = [(4, 1), (4, 3), (4, 5), (4, 7)]
    for (row, col), (label, formula, color) in zip(positions, kpis):
        ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col + 1)
        ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 3, end_column=col + 1)
        for r in range(row, row + 4):
            for c in range(col, col + 2):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=C_BG)
                ws.cell(r, c).border = BORDER
        ws.cell(row, col, label)
        ws.cell(row, col).font = Font(size=10, color=C_GRAY, name="Segoe UI")
        ws.cell(row + 2, col, value=formula)
        sty_kpi(ws.cell(row + 2, col), 18)
        ws.cell(row + 2, col).number_format = "#,##0"

    # Ventas por vendedor mini table
    ws.cell(10, 1, value="Ventas por vendedor (Mayo)")
    ws.cell(10, 1).font = Font(bold=True, name="Segoe UI", color=C_DARK)
    sty_header(ws.cell(11, 1, "Vendedor"))
    sty_header(ws.cell(11, 2, "Ventas S/"))
    vends = ["Ana Torres", "Luis Mendoza", "Carla Ríos", "Pedro Sánchez"]
    for i, v in enumerate(vends, 12):
        ws.cell(i, 1, v)
        ws.cell(i, 2, value=f'=SUMAR.SI.CONJUNTO(tblVentas[Monto S/];tblVentas[Vendedor];A{i};tblVentas[Mes];"Mayo")')
        ws.cell(i, 2).number_format = "#,##0"

    # Chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Ventas Mayo por vendedor"
    chart.style = 10
    chart.y_axis.title = "S/ "
    data = Reference(ws, min_col=2, min_row=11, max_row=15)
    cats = Reference(ws, min_col=1, min_row=12, max_row=15)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 14
    chart.height = 8
    ws.add_chart(chart, "D10")

    # Semáforo fill rate simulado
    ws.cell(10, 8, value="Fill rate sucursales (meta 95%)")
    ws.cell(10, 8).font = Font(bold=True, name="Segoe UI")
    fill_data = [("Trujillo", 0.82), ("Huancayo", 0.78), ("Lima", 0.96), ("Arequipa", 0.91)]
    sty_header(ws.cell(11, 8, "Sucursal"))
    sty_header(ws.cell(11, 9, "Fill %"))
    for i, (suc, pct) in enumerate(fill_data, 12):
        ws.cell(i, 8, suc)
        ws.cell(i, 9, pct)
        ws.cell(i, 9).number_format = "0%"
    ws.conditional_formatting.add(
        "I12:I15",
        FormulaRule(formula=["I12<0.85"], fill=PatternFill("solid", fgColor=C_NEG), font=Font(color="FFFFFF")),
    )
    ws.conditional_formatting.add(
        "I12:I15",
        FormulaRule(formula=["I12>=0.95"], fill=PatternFill("solid", fgColor=C_POS), font=Font(color="FFFFFF")),
    )

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 10


def build_simulador_meta(wb):
    ws = wb.create_sheet("12_Simulador_Meta")
    ws.sheet_properties.tabColor = "F1C40F"
    sty_title(ws, "Módulo 7 · Buscar Objetivo — modelo de utilidad")

    ws.cell(4, 2, "Unidades a vender")
    ws.cell(4, 3, 8500)
    ws.cell(5, 2, "Precio unitario S/")
    ws.cell(5, 3, 28.50)
    ws.cell(6, 2, "Costo variable S/")
    ws.cell(6, 3, 22.00)
    ws.cell(7, 2, "Costos fijos S/")
    ws.cell(7, 3, 95000)
    ws.cell(8, 2, "UTILIDAD S/")
    ws.cell(8, 2).font = Font(bold=True, size=12)
    ws.cell(8, 3, value="=C4*(C5-C6)-C7")
    ws.cell(8, 3).number_format = "#,##0"
    ws.cell(8, 3).font = Font(bold=True, size=14, color=C_DARK)

    ws.cell(10, 2, "Meta utilidad S/")
    ws.cell(10, 3, 400000)
    ws.cell(10, 3).fill = PatternFill("solid", fgColor=C_WARN)

    ws.cell(12, 2, value="→ Datos → Buscar objetivo → Celda C8 = 400000 → Cambiar C4")
    ws.cell(12, 2).font = Font(color=C_PRIMARY, name="Segoe UI", italic=True)

    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16


def build_escenarios(wb):
    ws = wb.create_sheet("13_Escenarios")
    ws.sheet_properties.tabColor = "F1C40F"
    sty_title(ws, "Módulo 7 · Administrador de Escenarios — variables y resultados")

    ws.cell(4, 2, "Variables (cambiar)")
    ws.cell(4, 2).font = Font(bold=True)
    labels = [
        ("Cambio dólar %", 0),
        ("Costo logístico %", 0),
        ("Pérdida ventas cliente S/", 0),
    ]
    for i, (lab, val) in enumerate(labels, 5):
        ws.cell(i, 2, lab)
        ws.cell(i, 3, val)

    ws.cell(9, 2, "Resultados")
    ws.cell(9, 2).font = Font(bold=True)
    ws.cell(10, 2, "Ventas base S/")
    ws.cell(10, 3, 4080000)
    ws.cell(11, 2, "Impacto variables S/")
    ws.cell(11, 3, value="=-C10*C5/100-C10*C6/100-C7")
    ws.cell(12, 2, "Ventas netas S/")
    ws.cell(12, 3, value="=C10+C11")
    ws.cell(13, 2, "Margen %")
    ws.cell(13, 3, 0.142)
    ws.cell(14, 2, "UTILIDAD S/")
    ws.cell(14, 3, value="=C12*C13")
    ws.cell(14, 3).number_format = "#,##0"
    ws.cell(14, 3).font = Font(bold=True, size=12)

    ws.cell(16, 2, value="Escenarios sugeridos: Base (0,0,0) | Dólar+5% | Logística+12% | Perdida 150000")
    ws.cell(16, 2).font = Font(italic=True, color=C_GRAY, name="Segoe UI")

    for r in range(10, 15):
        ws.cell(r, 3).number_format = "#,##0"

    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 16


def build_solver(wb):
    ws = wb.create_sheet("14_Solver_Promociones")
    ws.sheet_properties.tabColor = "F1C40F"
    sty_title(ws, "Módulo 7 · Solver — maximizar utilidad promocional (S/ 200K presupuesto)")

    ws.cell(4, 1, "Categoría")
    ws.cell(4, 2, "Presupuesto S/")
    ws.cell(4, 3, "Retorno marginal")
    ws.cell(4, 4, "Utilidad incremental")
    for i, h in enumerate(range(1, 5), 1):
        sty_header(ws.cell(4, i))

    cats = [
        ("Abarrotes", 0, 3.2),
        ("Lácteos", 0, 2.1),
        ("Bebidas", 0, 2.8),
        ("Limpieza", 0, 1.9),
        ("Snacks", 0, 3.5),
    ]
    for r, (cat, pres, ret) in enumerate(cats, 5):
        ws.cell(r, 1, cat)
        ws.cell(r, 2, pres)
        ws.cell(r, 2).number_format = "#,##0"
        ws.cell(r, 3, ret)
        ws.cell(r, 4, value=f"=B{r}*C{r}")
        ws.cell(r, 4).number_format = "#,##0"

    ws.cell(10, 1, "TOTAL PRESUPUESTO")
    ws.cell(10, 2, value="=SUMA(B5:B9)")
    ws.cell(10, 2).number_format = "#,##0"
    ws.cell(11, 1, "UTILIDAD TOTAL INCREMENTAL")
    ws.cell(11, 4, value="=SUMA(D5:D9)")
    ws.cell(11, 4).number_format = "#,##0"
    ws.cell(11, 4).font = Font(bold=True, size=12, color=C_POS)

    ws.cell(13, 1, value="Solver: Maximizar D11 | Variables B5:B9 | Restricción B10=200000 | B5:B9<=60000 | B5:B9>=0")
    ws.cell(13, 1).font = Font(color=C_PRIMARY, name="Segoe UI")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18


def build_vba_sheet(wb):
    ws = wb.create_sheet("15_Automatizacion_VBA")
    ws.sheet_properties.tabColor = "7F8C8D"
    sty_title(ws, "Módulo 6 · Código VBA de referencia — copiar al Editor (Alt+F11)")

    code = """
Sub ReporteDiario_Andina()
    Application.ScreenUpdating = False
    Application.EnableEvents = False
    On Error GoTo ErrorHandler

    ThisWorkbook.RefreshAll
    Application.CalculateUntilAsyncQueriesDone

    Dim ws As Worksheet
    Dim rutaPDF As String
    ws = ThisWorkbook.Worksheets("11_Dashboard_KPIs")
    rutaPDF = "C:\\Reportes\\PDF\\Ventas_" & Format(Date, "yyyymmdd") & ".pdf"
    ws.ExportAsFixedFormat Type:=xlTypePDF, Filename:=rutaPDF, Quality:=xlQualityStandard

    Application.EnableEvents = True
    Application.ScreenUpdating = True
    MsgBox "Reporte generado: " & rutaPDF, vbInformation
    Exit Sub
ErrorHandler:
    Application.EnableEvents = True
    Application.ScreenUpdating = True
    MsgBox "Error: " & Err.Description, vbCritical
End Sub
"""
    ws.cell(3, 1, value=code.strip())
    ws.cell(3, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(3, 1).font = Font(name="Consolas", size=9)
    ws.merge_cells("A3:F25")
    ws.row_dimensions[3].height = 280

    ws.cell(27, 1, value="Asignar macro a botón: Insertar → Formas → asignar ReporteDiario_Andina")
    ws.cell(27, 1).font = Font(color=C_PRIMARY, name="Segoe UI")


def build_contenido_curso(wb):
    ws = wb.create_sheet("17_Mapas_Contenido_Curso")
    ws.sheet_properties.tabColor = C_PRIMARY
    sty_title(ws, "Contenido del curso — 22 sub-temas (resumen ejecutivo + hoja demo)")

    headers = ["Módulo", "Sub-tema", "Gancho empresarial", "Fórmula / herramienta clave", "Hoja demo"]
    for i, h in enumerate(headers, 1):
        sty_header(ws.cell(3, i))

    rows = [
        ("1", "Formatos condicionales con fórmulas", "Alertas visuales sin revisar 500 filas manualmente", "=Y(G2<F2;G2<999) en fila completa", "01_Inventario_Alertas"),
        ("1", "Validación dependiente", "Región→Sucursal coherente, cero tipeos en formularios", "=INDIRECTO(C6) en validación lista", "04_Formulario_Cascada"),
        ("1", "Herramientas de datos", "Duplicados y consolidación sin inflar revenue", "Datos→Quitar duplicados / Consolidar", "10_Datos_Pivot_Ventas"),
        ("1", "Tablas Oficiales", "Fórmulas que se expanden al pegar filas nuevas", "=SUMA(tblVentas[Monto])", "01_Inventario_Alertas"),
        ("2", "BUSCARX", "Cruce catálogos cualquier dirección, sin #N/A", "=BUSCARX(RUC;tblClientes[RUC];tblClientes[Segmento])", "05_Ventas_Cruce"),
        ("2", "ÍNDICE + COINCIDIR", "Búsqueda 2D y tablas de comisiones escalonadas", "=ÍNDICE(tabla;COINCIDIR(val;col;0);COINCIDIR(val2;row;0))", "08_Liquidacion_SUMIFS"),
        ("2", "SI + Y + O anidados", "Clasificación automática Riesgo/Estrella/Estable", "=SI.CONJUNTO(...)", "09_Cartera_Mora"),
        ("2", "SUMIFS / COUNTIFS", "¿Ventas vendedor X en mayo? en 8 segundos", "=SUMAR.SI.CONJUNTO(tblVentas[Monto];...)", "08_Liquidacion_SUMIFS"),
        ("2", "DIAS.LAB.INTL", "Mora real sin fines de semana ni feriados", "=DIAS.LAB.INTL(venc;HOY();11;tblFeriados)", "09_Cartera_Mora"),
        ("2", "Texto avanzado", "Limpiar exports ERP sin pedir cambio al IT", "EXTRAER, ENCONTRAR, SUSTITUIR", "05_Ventas_Cruce"),
        ("3", "Campos calculados pivot", "Margen S/ dentro del pivot refrescable", "=Monto*Margen%/100 en campo calculado", "10_Datos_Pivot_Ventas"),
        ("3", "Segmentación horizontal", "Un slicer cambia todas las pivots del tablero", "Conexiones de informe en slicer", "11_Dashboard_KPIs"),
        ("3", "Agrupaciones avanzadas", "Trimestre con drill-down a mes en pivot", "Clic derecho Fecha→Agrupar", "10_Datos_Pivot_Ventas"),
        ("4", "Modelado de datos", "Cruce 3 tablas sin BUSCARV masivo", "Modelo estrella + relaciones", "05_Ventas + 06 + 07"),
        ("4", "ETL carpeta", "12 CSV combinados con Actualizar todo", "PQ: Desde carpeta→Combinar", "10_Datos_Pivot_Ventas"),
        ("4", ">1M filas", "Análisis sin truncar muestra", "Cargar a modelo, no a hoja", "— ver guía M4 en markdown"),
        ("4", "DAX básico", "KPIs que responden a slicers al instante", "Ventas Total = SUM(Ventas[Monto])", "11_Dashboard_KPIs"),
        ("5", "UI/UX corporativo", "GG decide en 3 min, no 45 min", "Sin gridlines + KPI cards + jerarquía", "11_Dashboard_KPIs"),
        ("5", "Paletas corporativas", "Misma identidad que Power BI y PPT", "Cian #00B4D8 / tema personalizado", "11_Dashboard_KPIs"),
        ("5", "Conexiones informe", "Gráfico y tabla muestran lo mismo", "Slicer→Conexiones de informe", "11_Dashboard_KPIs"),
        ("5", "Data Storytelling", "SCR: situación, conflicto, resolución", "Anotaciones + 3 bullets acción", "11_Dashboard_KPIs"),
        ("6", "VBA limpio", "Macro que no rompe al agregar columna", "ListObject sin .Select", "15_Automatizacion_VBA"),
        ("6", "Reporte diario auto", "90 min manual → 2 min RefreshAll+PDF", "ReporteDiario_Andina macro", "15_Automatizacion_VBA"),
        ("7", "Buscar objetivo", "¿Cuántas unidades para S/ 400K utilidad?", "Datos→Buscar objetivo en C8", "12_Simulador_Meta"),
        ("7", "Administrador escenarios", "Base vs Dólar+5% vs Crisis logística", "Datos→Administrador escenarios", "13_Escenarios"),
        ("7", "Solver", "Optimizar S/ 200K promo bajo restricciones", "Max D11 | B10=200000 | B<=60000", "14_Solver_Promociones"),
    ]
    for r, row in enumerate(rows, 4):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.font = Font(name="Segoe UI", size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER
        if row[4] != "— ver guía M4 en markdown":
            ws.cell(r, 5).hyperlink = f"#'{row[4]}'!A1"
            ws.cell(r, 5).font = Font(color=C_PRIMARY, underline="single", size=9)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["E"].width = 22
    for r in range(4, 4 + len(rows)):
        ws.row_dimensions[r].height = 36

    ws.cell(4 + len(rows) + 1, 1, value="Curso completo en Markdown: content/excel/Curso_Avanzado_Gestion_Empresarial_Excel.md")
    ws.cell(4 + len(rows) + 1, 1).font = Font(color=C_GRAY, name="Segoe UI", size=9)


def build_guia(wb):
    ws = wb.create_sheet("16_Guia_Rapida")
    ws.sheet_properties.tabColor = C_PRIMARY
    sty_title(ws, "Guía rápida — 7 módulos del curso")

    modulos = [
        ("M1", "Fundamentos y datos", "Tablas oficiales, formato condicional, validación cascada, herramientas datos"),
        ("M2", "Fórmulas avanzadas", "BUSCARX, ÍNDICE+COINCIDIR, SUMIFS, DIAS.LAB.INTL, SI.CONJUNTO, texto"),
        ("M3", "Tablas dinámicas", "Campos calculados, slicers conectados, agrupaciones — usar hoja 10_Datos_Pivot"),
        ("M4", "Power Query / Pivot", "Combinar carpeta CSV, modelo estrella, DAX — extender con Datos→Obtener datos"),
        ("M5", "Dashboards", "UI corporativa, paletas, storytelling — ver hoja 11_Dashboard_KPIs"),
        ("M6", "Macros VBA", "Código limpio, RefreshAll + PDF — hoja 15_Automatizacion_VBA"),
        ("M7", "Escenarios", "Buscar objetivo (12), Escenarios (13), Solver (14)"),
        ("Mapa", "22 sub-temas", "Ver hoja 17_Mapas_Contenido_Curso — ganchos, fórmulas y demos"),
    ]
    sty_header(ws.cell(3, 1, "Mód"))
    sty_header(ws.cell(3, 2, "Tema"))
    sty_header(ws.cell(3, 3, "Contenido"))
    for r, (m, t, c) in enumerate(modulos, 4):
        ws.cell(r, 1, m)
        ws.cell(r, 2, t)
        ws.cell(r, 3, c)
        for col in range(1, 4):
            ws.cell(r, col).font = Font(name="Segoe UI", size=10)
            ws.cell(r, col).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60


def main():
    wb = Workbook()
    build_index(wb)
    build_inventario(wb)
    build_facturas(wb)
    build_catalogo(wb)
    build_formulario(wb)
    build_maestro_clientes(wb)
    build_maestro_productos(wb)
    build_ventas(wb)
    build_liquidacion(wb)
    build_cartera(wb)
    build_pivot_data(wb)
    build_dashboard(wb)
    build_simulador_meta(wb)
    build_escenarios(wb)
    build_solver(wb)
    build_vba_sheet(wb)
    build_guia(wb)
    build_contenido_curso(wb)

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    wb.save(OUTPUT)
    print(f"Generado: {OUTPUT}")


if __name__ == "__main__":
    main()
